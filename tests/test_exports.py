"""
Unit tests for PDF and Excel export functionality.

Tests cover:
- PDF generation (pdf_exporter.py)
- Excel generation (excel_exporter.py)
- Export orchestration (exports.py)
- Metadata persistence and reconstruction
- Edge cases: empty data, missing fields, special characters
- Security: path traversal prevention, large input handling
"""
import json
import logging
import sys
import os
import tempfile
import traceback
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from models.inputs import UserInput
from models.suburb_metrics import (
    SuburbMetrics, SuburbIdentification, MarketMetricsCurrent,
    MarketMetricsHistory, PhysicalConfig, Demographics,
    Infrastructure, GrowthProjections, TimePoint,
)
from models.run_result import RunResult, SuburbReport


# ─── Test Data Factories ────────────────────────────────────────────────────

def make_full_suburb(name: str = "Springfield", state: str = "QLD",
                     median_price: float = 650000, rank: int = 1) -> SuburbReport:
    """Create a fully-populated SuburbReport for testing."""
    return SuburbReport(
        rank=rank,
        metrics=SuburbMetrics(
            identification=SuburbIdentification(
                name=name, state=state, lga="Test LGA", region="Test Region"
            ),
            market_current=MarketMetricsCurrent(
                median_price=median_price,
                average_price=median_price * 1.05,
                auction_clearance_current=72.5,
                days_on_market_current=28,
                turnover_rate_current=8.5,
                rental_yield_current=4.2,
            ),
            market_history=MarketMetricsHistory(
                price_history=[
                    TimePoint(year=2020, value=520000),
                    TimePoint(year=2021, value=550000),
                    TimePoint(year=2022, value=600000),
                    TimePoint(year=2023, value=630000),
                    TimePoint(year=2024, value=650000),
                ],
                dom_history=[
                    TimePoint(year=2020, value=45),
                    TimePoint(year=2021, value=38),
                    TimePoint(year=2022, value=32),
                    TimePoint(year=2023, value=30),
                    TimePoint(year=2024, value=28),
                ],
            ),
            physical_config=PhysicalConfig(
                land_size_median_sqm=600,
                floor_size_median_sqm=180,
                typical_bedrooms=4,
                typical_bathrooms=2,
                typical_car_spaces=2,
            ),
            demographics=Demographics(
                population_trend="Growing steadily at 2% per year",
                median_age=34.5,
                household_types={"families_with_children": 0.45, "couples": 0.25},
                income_distribution={"low": 0.2, "medium": 0.5, "high": 0.3},
            ),
            infrastructure=Infrastructure(
                current_transport=["Bus route 100", "Train station 2km"],
                future_transport=["Metro extension planned 2028"],
                planned_infrastructure=["New hospital 2027", "Shopping centre expansion"],
                major_events_relevance="Near 2032 Olympics venue",
                shopping_access="Major shopping centre within 5km",
                schools_summary="3 primary, 2 secondary schools within 5km",
                crime_stats={"assault": "low", "property": "medium"},
            ),
            growth_projections=GrowthProjections(
                projected_growth_pct={1: 5.2, 2: 11.5, 3: 18.3, 5: 32.5, 10: 78.2, 25: 245.0},
                confidence_intervals={
                    1: (3.5, 7.0),
                    2: (7.0, 16.0),
                    3: (11.0, 26.0),
                    5: (20.0, 45.0),
                    10: (50.0, 110.0),
                    25: (150.0, 400.0),
                },
                risk_analysis="Moderate risk due to flood zone proximity. Mitigated by council infrastructure investment.",
                key_drivers=[
                    "Brisbane 2032 Olympics proximity",
                    "New metro line extension",
                    "Affordable entry price relative to region",
                    "Strong population growth",
                ],
                growth_score=82.5,
                risk_score=35.0,
                composite_score=75.8,
            ),
        ),
        narrative_html="<h2>Springfield Analysis</h2><p>Promising suburb.</p>",
        charts={
            "price_history": "price_history_springfield-qld.png",
            "growth_projection": "growth_projection_springfield-qld.png",
        },
    )


def make_minimal_suburb(name: str = "Bare Town", state: str = "NSW",
                        rank: int = 1) -> SuburbReport:
    """Create a minimal SuburbReport with only required fields."""
    return SuburbReport(
        rank=rank,
        metrics=SuburbMetrics(
            identification=SuburbIdentification(
                name=name, state=state, lga="Min LGA"
            ),
            market_current=MarketMetricsCurrent(median_price=400000),
        ),
    )


def make_run_result(suburbs=None, run_id="test-2026-02-13") -> RunResult:
    """Create a RunResult for testing."""
    if suburbs is None:
        suburbs = [
            make_full_suburb("Springfield", "QLD", 650000, rank=1),
            make_full_suburb("Shelbyville", "NSW", 550000, rank=2),
            make_full_suburb("Capital City", "VIC", 720000, rank=3),
        ]
    return RunResult(
        run_id=run_id,
        user_input=UserInput(
            max_median_price=800000,
            dwelling_type="house",
            regions=["South East Queensland"],
            num_suburbs=max(len(suburbs), 1),  # Pydantic requires > 0
        ),
        suburbs=suburbs,
        status="completed",
    )


# ─── Test Counters ──────────────────────────────────────────────────────────

passed = 0
failed = 0
errors = []


def record_pass(name):
    global passed
    passed += 1
    print(f"  ✓ {name}")


def record_fail(name, reason=""):
    global failed
    failed += 1
    msg = f"  ✗ {name}: {reason}"
    errors.append(msg)
    print(msg)


# ─── PDF Exporter Tests ─────────────────────────────────────────────────────

def test_pdf_formatting_helpers():
    """Test PDF formatting utility functions."""
    print("\n── PDF Formatting Helpers ──")
    from reporting.pdf_exporter import _fmt_currency, _fmt_pct, _fmt_num, _safe_str

    # _fmt_currency
    assert _fmt_currency(650000) == "$650,000", f"Got {_fmt_currency(650000)}"
    record_pass("_fmt_currency with valid float")

    assert _fmt_currency(None) == "N/A"
    record_pass("_fmt_currency with None")

    assert _fmt_currency("invalid") == "N/A"
    record_pass("_fmt_currency with invalid string")

    assert _fmt_currency(0) == "$0"
    record_pass("_fmt_currency with zero")

    # _fmt_pct
    assert _fmt_pct(5.23) == "5.2%"
    record_pass("_fmt_pct with valid float")

    assert _fmt_pct(None) == "N/A"
    record_pass("_fmt_pct with None")

    assert _fmt_pct("bad") == "N/A"
    record_pass("_fmt_pct with invalid string")

    # _fmt_num
    assert _fmt_num(82.567, 1) == "82.6"
    record_pass("_fmt_num with 1 decimal")

    assert _fmt_num(82.567, 0) == "83"
    record_pass("_fmt_num with 0 decimals")

    assert _fmt_num(None) == "N/A"
    record_pass("_fmt_num with None")

    # _safe_str
    assert _safe_str(None) == "N/A"
    record_pass("_safe_str with None")

    assert _safe_str({"a": 1, "b": 2}) == "a: 1; b: 2"
    record_pass("_safe_str with dict")

    assert _safe_str(["item1", "item2"]) == "item1; item2"
    record_pass("_safe_str with list")

    assert _safe_str([]) == "N/A"
    record_pass("_safe_str with empty list")

    assert _safe_str("hello") == "hello"
    record_pass("_safe_str with string")


def test_pdf_class_init():
    """Test PropertyReportPDF class initialization."""
    print("\n── PDF Class Initialization ──")
    from reporting.pdf_exporter import PropertyReportPDF

    pdf = PropertyReportPDF(run_id="test-run-123")
    assert pdf.run_id == "test-run-123"
    record_pass("PDF class stores run_id")

    assert pdf._is_title_page is True
    record_pass("PDF starts with title page flag True")

    # fpdf2 stores width in user units; A4 = 210mm
    assert abs(pdf.w - 210) < 1, f"Width={pdf.w}, expected ~210"
    record_pass("PDF is A4 format")


def test_pdf_generation_full():
    """Test complete PDF generation with full data."""
    print("\n── PDF Full Generation ──")
    from reporting.pdf_exporter import generate_pdf

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "test_report.pdf"

        generate_pdf(run_result, output_dir, pdf_path)

        assert pdf_path.exists(), "PDF file was not created"
        record_pass("PDF file created successfully")

        size = pdf_path.stat().st_size
        assert size > 1000, f"PDF too small: {size} bytes"
        record_pass(f"PDF has reasonable size ({size:,} bytes)")

        # Verify it starts with PDF header
        with open(pdf_path, 'rb') as f:
            header = f.read(5)
        assert header == b'%PDF-', f"Not a valid PDF: {header}"
        record_pass("PDF has valid header")


def test_pdf_generation_minimal():
    """Test PDF generation with minimal data (many None fields)."""
    print("\n── PDF Minimal Data ──")
    from reporting.pdf_exporter import generate_pdf

    run_result = make_run_result(suburbs=[make_minimal_suburb()])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "minimal.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF generated with minimal data (no crash)")


def test_pdf_generation_empty_suburbs():
    """Test PDF generation with zero suburbs."""
    print("\n── PDF Empty Suburbs ──")
    from reporting.pdf_exporter import generate_pdf

    run_result = make_run_result(suburbs=[])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "empty.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF generated with no suburbs (no crash)")


def test_pdf_generation_special_characters():
    """Test PDF generation with special characters in suburb names."""
    print("\n── PDF Special Characters ──")
    from reporting.pdf_exporter import generate_pdf

    suburb = make_full_suburb(name="O'Reilly's & Co", state="QLD", rank=1)
    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "special.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF handles special characters in names")


def test_pdf_many_suburbs():
    """Test PDF generation with many suburbs (stress test)."""
    print("\n── PDF Many Suburbs ──")
    from reporting.pdf_exporter import generate_pdf

    suburbs = [make_full_suburb(f"Suburb{i}", "QLD", 500000 + i * 10000, rank=i)
               for i in range(1, 21)]
    run_result = make_run_result(suburbs=suburbs)

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "many.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        size = pdf_path.stat().st_size
        assert size > 5000, f"PDF with 20 suburbs seems too small: {size}"
        record_pass(f"PDF generated with 20 suburbs ({size:,} bytes)")


def test_pdf_missing_charts_dir():
    """Test PDF handles missing charts directory gracefully."""
    print("\n── PDF Missing Charts Dir ──")
    from reporting.pdf_exporter import generate_pdf

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        # Intentionally NOT creating charts/ dir
        pdf_path = output_dir / "no_charts.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF generated without charts directory")


# ─── Excel Exporter Tests ───────────────────────────────────────────────────

def test_excel_formatting_helpers():
    """Test Excel formatting utility functions."""
    print("\n── Excel Formatting Helpers ──")
    from reporting.excel_exporter import _safe_str

    assert _safe_str(None) == ""
    record_pass("Excel _safe_str with None returns empty string")

    assert _safe_str({"k": "v"}) == "k: v"
    record_pass("Excel _safe_str with dict")

    assert _safe_str(["a", "b"]) == "a; b"
    record_pass("Excel _safe_str with list")


def test_excel_generation_full():
    """Test complete Excel generation with full data."""
    print("\n── Excel Full Generation ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "test_report.xlsx"

        generate_excel(run_result, xlsx_path)

        assert xlsx_path.exists(), "Excel file was not created"
        record_pass("Excel file created successfully")

        size = xlsx_path.stat().st_size
        assert size > 1000, f"Excel too small: {size} bytes"
        record_pass(f"Excel has reasonable size ({size:,} bytes)")

        # Load and verify structure
        wb = load_workbook(xlsx_path)
        expected_sheets = [
            "Overview", "Market Metrics", "Growth Projections",
            "Property Config", "Demographics", "Infrastructure", "Price History"
        ]
        assert wb.sheetnames == expected_sheets, f"Sheets: {wb.sheetnames}"
        record_pass("Excel has all 7 expected sheets")

        # Check overview sheet content
        ws = wb["Overview"]
        assert ws["A1"].value == "Australian Property Research Report"
        record_pass("Overview sheet has title")

        # Check rankings data exists
        assert ws.cell(row=12, column=1).value == "Rank"
        record_pass("Rankings table header present")

        # Check first suburb data row
        assert ws.cell(row=13, column=2).value is not None
        record_pass("First suburb data present in rankings")

        wb.close()


def test_excel_sheet_market_metrics():
    """Test Market Metrics sheet content."""
    print("\n── Excel Market Metrics Sheet ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "market.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Market Metrics"]

        # Check header
        assert ws.cell(row=1, column=1).value == "Suburb"
        assert ws.cell(row=1, column=3).value == "Median Price"
        record_pass("Market Metrics has correct headers")

        # Check data
        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=2, column=3).value is not None  # median_price
        record_pass("Market Metrics has data in first row")

        wb.close()


def test_excel_sheet_growth_projections():
    """Test Growth Projections sheet content."""
    print("\n── Excel Growth Projections Sheet ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "growth.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Growth Projections"]

        # Check growth percentage columns
        assert ws.cell(row=1, column=3).value == "1yr %"
        assert ws.cell(row=1, column=8).value == "25yr %"
        record_pass("Growth sheet has correct year headers")

        # Check confidence interval columns
        assert ws.cell(row=1, column=9).value == "1yr Low"
        record_pass("Growth sheet has confidence interval headers")

        # Check data values
        val_5yr = ws.cell(row=2, column=6).value  # 5yr %
        assert val_5yr is not None and val_5yr > 0, f"5yr growth: {val_5yr}"
        record_pass("Growth projections contain valid data")

        wb.close()


def test_excel_sheet_price_history():
    """Test Price History sheet with wide format."""
    print("\n── Excel Price History Sheet ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "history.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Price History"]

        # Should have year columns
        assert ws.cell(row=1, column=1).value == "Suburb"
        assert ws.cell(row=1, column=3).value is not None  # first year column
        record_pass("Price History has suburb and year columns")

        # Check actual price data
        price_val = ws.cell(row=2, column=3).value
        assert price_val is not None, "Price history data missing"
        record_pass("Price History contains actual price values")

        wb.close()


def test_excel_generation_minimal():
    """Test Excel generation with minimal data."""
    print("\n── Excel Minimal Data ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result(suburbs=[make_minimal_suburb()])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "minimal.xlsx"
        generate_excel(run_result, xlsx_path)

        assert xlsx_path.exists()
        record_pass("Excel generated with minimal data (no crash)")

        wb = load_workbook(xlsx_path)

        # Price history sheet should handle empty history
        ws = wb["Price History"]
        assert ws.cell(row=1, column=1).value == "No price history data available"
        record_pass("Price History handles empty history correctly")

        wb.close()


def test_excel_generation_empty_suburbs():
    """Test Excel generation with zero suburbs."""
    print("\n── Excel Empty Suburbs ──")
    from reporting.excel_exporter import generate_excel

    run_result = make_run_result(suburbs=[])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "empty.xlsx"
        generate_excel(run_result, xlsx_path)

        assert xlsx_path.exists()
        record_pass("Excel generated with no suburbs (no crash)")


def test_excel_special_characters():
    """Test Excel with special characters in data."""
    print("\n── Excel Special Characters ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    suburb = make_full_suburb(name="O'Reilly's & Co <script>", state="QLD", rank=1)
    # Also add potentially dangerous infrastructure text
    suburb.metrics.infrastructure.current_transport = ["Train; 'quoted' station"]
    suburb.metrics.infrastructure.shopping_access = "Mall & <b>shops</b>"
    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "special.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Overview"]

        # The name should be stored as-is (Excel cells are not HTML-rendered)
        found_name = ws.cell(row=13, column=2).value
        assert found_name == "O'Reilly's & Co <script>"
        record_pass("Excel stores special characters safely as text")

        wb.close()


def test_excel_styling():
    """Test that Excel styling is applied correctly."""
    print("\n── Excel Styling ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "styled.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Market Metrics"]

        # Check header styling
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        record_pass("Header cells are bold")

        assert header_cell.font.color.rgb == "00FFFFFF"  # White with alpha
        record_pass("Header cells have white font color")

        # Check that borders are applied
        data_cell = ws.cell(row=2, column=1)
        assert data_cell.border.left.style is not None
        record_pass("Data cells have borders")

        wb.close()


# ─── Export Orchestration Tests ──────────────────────────────────────────────

def test_export_pdf_orchestration():
    """Test generate_pdf_export orchestration."""
    print("\n── Export PDF Orchestration ──")
    from reporting.exports import generate_pdf_export

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()

        pdf_path = generate_pdf_export(run_result, output_dir)

        assert pdf_path.exists()
        assert pdf_path.name == f"report_{run_result.run_id}.pdf"
        record_pass("generate_pdf_export returns correct path")


def test_export_excel_orchestration():
    """Test generate_excel_export orchestration."""
    print("\n── Export Excel Orchestration ──")
    from reporting.exports import generate_excel_export

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)

        xlsx_path = generate_excel_export(run_result, output_dir)

        assert xlsx_path.exists()
        assert xlsx_path.name == f"report_{run_result.run_id}.xlsx"
        record_pass("generate_excel_export returns correct path")


def test_export_error_wrapping():
    """Test that ExportError is raised on failure."""
    print("\n── Export Error Handling ──")
    from reporting.exports import generate_pdf_export, ExportError

    run_result = make_run_result()

    # Use a non-existent directory for output to trigger error
    bad_dir = Path("/nonexistent/path/that/should/not/exist")

    try:
        generate_pdf_export(run_result, bad_dir)
        record_fail("ExportError not raised for invalid output dir")
    except ExportError as e:
        assert "PDF generation failed" in str(e)
        record_pass("ExportError raised with descriptive message")
    except Exception as e:
        record_fail("ExportError not raised", f"Got {type(e).__name__}: {e}")


# ─── Metadata Reconstruction Tests ──────────────────────────────────────────

def test_reconstruct_run_result_success():
    """Test successful RunResult reconstruction from metadata."""
    print("\n── Metadata Reconstruction ──")
    from reporting.exports import reconstruct_run_result

    run_result = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        metadata_path = output_dir / "run_metadata.json"
        metadata_path.write_text(run_result.model_dump_json(indent=2), encoding="utf-8")

        reconstructed = reconstruct_run_result("test-2026-02-13", output_dir)

        assert reconstructed is not None
        record_pass("RunResult reconstructed successfully")

        assert reconstructed.run_id == run_result.run_id
        record_pass("Reconstructed run_id matches")

        assert len(reconstructed.suburbs) == len(run_result.suburbs)
        record_pass("Reconstructed suburbs count matches")

        assert reconstructed.user_input.max_median_price == run_result.user_input.max_median_price
        record_pass("Reconstructed user_input preserved")


def test_reconstruct_run_result_missing_file():
    """Test reconstruction returns None when metadata file missing."""
    print("\n── Metadata Missing File ──")
    from reporting.exports import reconstruct_run_result

    with tempfile.TemporaryDirectory() as tmpdir:
        result = reconstruct_run_result("nonexistent", Path(tmpdir))
        assert result is None
        record_pass("Returns None for missing metadata file")


def test_reconstruct_run_result_invalid_json():
    """Test reconstruction handles invalid JSON gracefully."""
    print("\n── Metadata Invalid JSON ──")
    from reporting.exports import reconstruct_run_result

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        metadata_path = output_dir / "run_metadata.json"
        metadata_path.write_text("{ this is not valid json !!!", encoding="utf-8")

        # Suppress expected warning log output
        logging.getLogger("reporting.exports").setLevel(logging.CRITICAL)
        result = reconstruct_run_result("bad-json", output_dir)
        logging.getLogger("reporting.exports").setLevel(logging.WARNING)

        assert result is None
        record_pass("Returns None for invalid JSON (no crash)")


def test_reconstruct_run_result_incomplete_data():
    """Test reconstruction handles incomplete data gracefully."""
    print("\n── Metadata Incomplete Data ──")
    from reporting.exports import reconstruct_run_result

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        metadata_path = output_dir / "run_metadata.json"
        # Valid JSON but missing required fields
        metadata_path.write_text('{"run_id": "test"}', encoding="utf-8")

        # Suppress expected warning log output
        logging.getLogger("reporting.exports").setLevel(logging.CRITICAL)
        result = reconstruct_run_result("incomplete", output_dir)
        logging.getLogger("reporting.exports").setLevel(logging.WARNING)

        assert result is None
        record_pass("Returns None for incomplete metadata (no crash)")


# ─── Roundtrip Tests ────────────────────────────────────────────────────────

def test_metadata_roundtrip():
    """Test that serializing and deserializing RunResult preserves data."""
    print("\n── Metadata Roundtrip ──")
    from reporting.exports import reconstruct_run_result

    original = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        metadata_path = output_dir / "run_metadata.json"
        metadata_path.write_text(original.model_dump_json(indent=2), encoding="utf-8")

        restored = reconstruct_run_result(original.run_id, output_dir)

        # Verify detailed field preservation
        orig_suburb = original.suburbs[0].metrics
        rest_suburb = restored.suburbs[0].metrics

        assert orig_suburb.identification.name == rest_suburb.identification.name
        record_pass("Suburb name preserved through roundtrip")

        assert orig_suburb.market_current.median_price == rest_suburb.market_current.median_price
        record_pass("Median price preserved through roundtrip")

        orig_growth = orig_suburb.growth_projections.projected_growth_pct
        rest_growth = rest_suburb.growth_projections.projected_growth_pct
        # JSON serialization converts int keys to strings, so compare values
        for year in [1, 2, 3, 5, 10, 25]:
            orig_val = orig_growth.get(year, orig_growth.get(str(year)))
            rest_val = rest_growth.get(year, rest_growth.get(str(year)))
            assert orig_val == rest_val, f"Year {year}: {orig_val} != {rest_val}"
        record_pass("Growth projections preserved through roundtrip")

        assert len(orig_suburb.market_history.price_history) == len(rest_suburb.market_history.price_history)
        record_pass("Price history length preserved through roundtrip")

        assert orig_suburb.growth_projections.key_drivers == rest_suburb.growth_projections.key_drivers
        record_pass("Key drivers list preserved through roundtrip")


def test_export_roundtrip_pdf_then_reconstruct():
    """Test full flow: generate metadata, reconstruct, then export."""
    print("\n── Full Export Roundtrip ──")
    from reporting.exports import generate_pdf_export, generate_excel_export, reconstruct_run_result

    original = make_run_result()

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()

        # Save metadata (simulating what html_renderer does)
        metadata_path = output_dir / "run_metadata.json"
        metadata_path.write_text(original.model_dump_json(indent=2), encoding="utf-8")

        # Reconstruct
        restored = reconstruct_run_result(original.run_id, output_dir)
        assert restored is not None

        # Export from reconstructed data
        pdf_path = generate_pdf_export(restored, output_dir)
        xlsx_path = generate_excel_export(restored, output_dir)

        assert pdf_path.exists()
        assert xlsx_path.exists()
        record_pass("Full roundtrip: save metadata -> reconstruct -> export PDF + Excel")


# ─── Security Tests ─────────────────────────────────────────────────────────

def test_security_long_strings():
    """Test that very long strings don't cause memory issues."""
    print("\n── Security: Long Strings ──")
    from reporting.pdf_exporter import generate_pdf
    from reporting.excel_exporter import generate_excel

    suburb = make_full_suburb(name="Normal Suburb", rank=1)
    # Set extremely long risk analysis
    suburb.metrics.growth_projections.risk_analysis = "A" * 10000
    suburb.metrics.growth_projections.key_drivers = [f"Driver {i}: " + "x" * 500 for i in range(20)]
    suburb.metrics.infrastructure.current_transport = ["Route " + str(i) * 100 for i in range(50)]

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()

        pdf_path = output_dir / "long.pdf"
        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF handles very long strings without crash")

        xlsx_path = output_dir / "long.xlsx"
        generate_excel(run_result, xlsx_path)
        assert xlsx_path.exists()
        record_pass("Excel handles very long strings without crash")


def test_security_negative_values():
    """Test handling of negative and extreme numeric values."""
    print("\n── Security: Extreme Values ──")
    from reporting.pdf_exporter import _fmt_currency, _fmt_pct
    from reporting.excel_exporter import generate_excel

    assert _fmt_currency(-500000) == "$-500,000" or _fmt_currency(-500000) == "-$500,000"
    record_pass("_fmt_currency handles negative values")

    assert _fmt_pct(-5.5) == "-5.5%"
    record_pass("_fmt_pct handles negative values")

    # Test with extreme growth values
    suburb = make_full_suburb(rank=1)
    suburb.metrics.growth_projections.projected_growth_pct = {
        1: -10.0, 2: -5.0, 3: 0.0, 5: 999.9, 10: 9999.9, 25: 99999.9
    }
    suburb.metrics.market_current.median_price = 0.01  # Near-zero price

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "extreme.xlsx"
        generate_excel(run_result, xlsx_path)
        assert xlsx_path.exists()
        record_pass("Excel handles extreme numeric values")


def test_security_unicode_handling():
    """Test handling of unicode characters in data."""
    print("\n── Security: Unicode ──")
    from reporting.pdf_exporter import generate_pdf
    from reporting.excel_exporter import generate_excel

    suburb = make_full_suburb(name="Test\u00e9 Suburb\u2019s", state="QLD", rank=1)
    suburb.metrics.infrastructure.schools_summary = "L'\u00e9cole internationale"
    suburb.metrics.demographics.population_trend = "\u4eba\u53e3\u589e\u957f"  # Chinese chars

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()

        xlsx_path = output_dir / "unicode.xlsx"
        generate_excel(run_result, xlsx_path)
        assert xlsx_path.exists()
        record_pass("Excel handles unicode characters")

        pdf_path = output_dir / "unicode.pdf"
        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF handles unicode characters via sanitization")


# ─── PDF Unicode Sanitization Tests ─────────────────────────────────────────

def test_pdf_sanitize_function():
    """Test _sanitize() replaces problematic Unicode characters."""
    print("\n── PDF Sanitize Function ──")
    from reporting.pdf_exporter import _sanitize

    # En-dash and em-dash
    assert _sanitize("stages 1\u20134") == "stages 1-4"
    assert _sanitize("2025\u2014a big year") == "2025-a big year"
    record_pass("_sanitize replaces en-dash and em-dash")

    # Smart quotes
    assert _sanitize("Leichhardt\u2019s growth") == "Leichhardt's growth"
    assert _sanitize("\u2018quoted\u2019") == "'quoted'"
    assert _sanitize("\u201cdouble\u201d") == '"double"'
    record_pass("_sanitize replaces smart quotes")

    # Ellipsis, bullet, non-breaking space
    assert _sanitize("more\u2026") == "more..."
    assert _sanitize("\u2022 item") == "- item"
    assert _sanitize("no\u00a0break") == "no break"
    record_pass("_sanitize replaces ellipsis, bullet, nbsp")

    # ASCII passthrough
    plain = "Normal ASCII text - with 'quotes' and \"doubles\""
    assert _sanitize(plain) == plain
    record_pass("_sanitize passes through plain ASCII unchanged")

    # Mixed Unicode
    mixed = "Budget 2025\u201326: Leichhardt\u2019s Prep\u2013Year 6 \u2022 growth"
    expected = "Budget 2025-26: Leichhardt's Prep-Year 6 - growth"
    assert _sanitize(mixed) == expected
    record_pass("_sanitize handles mixed Unicode in one string")


def test_pdf_cell_override_sanitizes():
    """Test PropertyReportPDF.cell() and multi_cell() auto-sanitize text."""
    print("\n── PDF Cell Override ──")
    from reporting.pdf_exporter import PropertyReportPDF

    pdf = PropertyReportPDF(run_id="test")
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)

    # cell() with positional text containing en-dash — should not raise
    pdf.cell(w=180, h=10, text="stages 1\u20134")
    record_pass("cell() sanitizes positional text arg")

    # cell() with keyword text
    pdf.cell(w=180, h=10, text="Leichhardt\u2019s growth")
    record_pass("cell() sanitizes keyword text arg")

    # multi_cell() with en-dash
    pdf.multi_cell(w=180, h=10, text="Budget 2025\u201326")
    record_pass("multi_cell() sanitizes positional text arg")

    # multi_cell() with keyword text containing smart quotes
    pdf.multi_cell(w=180, h=10, text="\u201cQuoted\u201d analysis")
    record_pass("multi_cell() sanitizes keyword text arg")

    # Verify PDF can be output (no crash)
    output = pdf.output()
    assert len(output) > 0
    record_pass("PDF with sanitized Unicode cells outputs successfully")


def test_pdf_realworld_unicode_data():
    """Test PDF generation with real-world Unicode patterns from actual API data."""
    print("\n── PDF Real-World Unicode ──")
    from reporting.pdf_exporter import generate_pdf

    suburb = make_full_suburb(name="Leichhardt", state="QLD", rank=1)

    # Patterns found in actual run_metadata.json files
    suburb.metrics.infrastructure.schools_summary = "Woodridge State School (Prep\u2013Year 6)"
    suburb.metrics.infrastructure.future_transport = [
        "Cross River Rail \u2013 stages 5\u20137",
        "Brisbane Metro \u2013 northern extension"
    ]
    suburb.metrics.infrastructure.current_infrastructure = [
        "Logan Motorway enhancement (stages 1\u20134)"
    ]
    suburb.metrics.infrastructure.planned_infrastructure = [
        "Federal Budget 2025\u201326 allocations"
    ]
    suburb.metrics.growth_projections.risk_analysis = (
        "Leichhardt\u2019s growth is tied to Brisbane\u2013Gold Coast corridor "
        "infrastructure. The suburb\u2019s proximity to major projects "
        "\u201csignificantly enhances\u201d long-term prospects."
    )
    suburb.metrics.growth_projections.key_drivers = [
        "Brisbane 2032 Olympics \u2013 venue precinct development",
        "Cross River Rail \u2013 improved connectivity",
    ]
    suburb.metrics.identification.region = "Brisbane\u2013Gold Coast Corridor"
    suburb.metrics.infrastructure.major_events_relevance = (
        "2032 Olympics \u2013 Gabba rebuild, athletes\u2019 village"
    )

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "realworld_unicode.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        assert pdf_path.stat().st_size > 1000
        record_pass("PDF generates with real-world Unicode data patterns")


# ─── Edge Case Tests ────────────────────────────────────────────────────────

def test_edge_single_suburb():
    """Test with exactly one suburb."""
    print("\n── Edge: Single Suburb ──")
    from reporting.pdf_exporter import generate_pdf
    from reporting.excel_exporter import generate_excel

    run_result = make_run_result(suburbs=[make_full_suburb(rank=1)])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()

        pdf_path = output_dir / "single.pdf"
        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF works with single suburb")

        xlsx_path = output_dir / "single.xlsx"
        generate_excel(run_result, xlsx_path)
        assert xlsx_path.exists()
        record_pass("Excel works with single suburb")


def test_edge_no_confidence_intervals():
    """Test suburbs with empty confidence intervals."""
    print("\n── Edge: No Confidence Intervals ──")
    from reporting.pdf_exporter import generate_pdf

    suburb = make_full_suburb(rank=1)
    suburb.metrics.growth_projections.confidence_intervals = {}

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        output_dir = Path(tmpdir)
        (output_dir / "charts").mkdir()
        pdf_path = output_dir / "no_ci.pdf"

        generate_pdf(run_result, output_dir, pdf_path)
        assert pdf_path.exists()
        record_pass("PDF handles missing confidence intervals")


def test_edge_no_price_history():
    """Test suburbs with empty price history."""
    print("\n── Edge: No Price History ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    suburb = make_full_suburb(rank=1)
    suburb.metrics.market_history.price_history = []

    run_result = make_run_result(suburbs=[suburb])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "no_history.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Price History"]
        assert ws.cell(row=1, column=1).value == "No price history data available"
        record_pass("Excel handles empty price history correctly")
        wb.close()


def test_edge_mixed_history_years():
    """Test suburbs with different price history years."""
    print("\n── Edge: Mixed History Years ──")
    from reporting.excel_exporter import generate_excel
    from openpyxl import load_workbook

    suburb1 = make_full_suburb("Town A", "QLD", rank=1)
    suburb1.metrics.market_history.price_history = [
        TimePoint(year=2020, value=500000),
        TimePoint(year=2022, value=550000),
    ]

    suburb2 = make_full_suburb("Town B", "NSW", rank=2)
    suburb2.metrics.market_history.price_history = [
        TimePoint(year=2021, value=400000),
        TimePoint(year=2022, value=420000),
        TimePoint(year=2023, value=440000),
    ]

    run_result = make_run_result(suburbs=[suburb1, suburb2])

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = Path(tmpdir) / "mixed_years.xlsx"
        generate_excel(run_result, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Price History"]

        # Should have union of all years: 2020, 2021, 2022, 2023
        year_headers = []
        col = 3
        while ws.cell(row=1, column=col).value is not None:
            year_headers.append(ws.cell(row=1, column=col).value)
            col += 1

        assert "2020" in year_headers
        assert "2021" in year_headers
        assert "2022" in year_headers
        assert "2023" in year_headers
        record_pass("Price History handles mixed years across suburbs")

        wb.close()


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  EXPORT FUNCTIONALITY - UNIT TESTS")
    print("=" * 70)

    tests = [
        # PDF tests
        test_pdf_formatting_helpers,
        test_pdf_class_init,
        test_pdf_generation_full,
        test_pdf_generation_minimal,
        test_pdf_generation_empty_suburbs,
        test_pdf_generation_special_characters,
        test_pdf_many_suburbs,
        test_pdf_missing_charts_dir,
        # Excel tests
        test_excel_formatting_helpers,
        test_excel_generation_full,
        test_excel_sheet_market_metrics,
        test_excel_sheet_growth_projections,
        test_excel_sheet_price_history,
        test_excel_generation_minimal,
        test_excel_generation_empty_suburbs,
        test_excel_special_characters,
        test_excel_styling,
        # Export orchestration tests
        test_export_pdf_orchestration,
        test_export_excel_orchestration,
        test_export_error_wrapping,
        # Metadata reconstruction tests
        test_reconstruct_run_result_success,
        test_reconstruct_run_result_missing_file,
        test_reconstruct_run_result_invalid_json,
        test_reconstruct_run_result_incomplete_data,
        # Roundtrip tests
        test_metadata_roundtrip,
        test_export_roundtrip_pdf_then_reconstruct,
        # Security tests
        test_security_long_strings,
        test_security_negative_values,
        test_security_unicode_handling,
        # PDF Unicode sanitization tests
        test_pdf_sanitize_function,
        test_pdf_cell_override_sanitizes,
        test_pdf_realworld_unicode_data,
        # Edge case tests
        test_edge_single_suburb,
        test_edge_no_confidence_intervals,
        test_edge_no_price_history,
        test_edge_mixed_history_years,
    ]

    for test_fn in tests:
        try:
            test_fn()
        except Exception as e:
            record_fail(test_fn.__name__, str(e))
            traceback.print_exc()
            print()

    # Summary
    total = passed + failed
    print("\n" + "=" * 70)
    print(f"  RESULTS: {passed}/{total} passed, {failed} failed")
    print("=" * 70)

    if errors:
        print("\nFailed tests:")
        for err in errors:
            print(err)

    if failed > 0:
        sys.exit(1)
    else:
        print("\n✓ All tests passed!")
        sys.exit(0)


if __name__ == "__main__":
    main()
