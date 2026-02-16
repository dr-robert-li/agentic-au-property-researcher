"""
Excel report generation using openpyxl.

Creates a multi-sheet .xlsx workbook with structured suburb research data.
"""
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.run_result import RunResult, SuburbReport

# Style constants
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(name='Calibri', bold=True, size=16, color='2E86AB')
LABEL_FONT = Font(name='Calibri', bold=True, size=11)
CURRENCY_FORMAT = '$#,##0'
PERCENT_FORMAT = '0.0'
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
ALT_ROW_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')


def generate_excel(run_result: RunResult, xlsx_path: Path):
    """
    Generate complete Excel workbook from run result data.

    Args:
        run_result: Complete RunResult with all suburb reports
        xlsx_path: Where to write the .xlsx file
    """
    wb = Workbook()

    # Sheet 1: Overview (default sheet)
    _create_overview_sheet(wb.active, run_result)

    # Sheet 2: Market Metrics
    _create_market_metrics_sheet(wb.create_sheet("Market Metrics"), run_result)

    # Sheet 3: Growth Projections
    _create_growth_sheet(wb.create_sheet("Growth Projections"), run_result)

    # Sheet 4: Property Config
    _create_property_config_sheet(wb.create_sheet("Property Config"), run_result)

    # Sheet 5: Demographics
    _create_demographics_sheet(wb.create_sheet("Demographics"), run_result)

    # Sheet 6: Infrastructure
    _create_infrastructure_sheet(wb.create_sheet("Infrastructure"), run_result)

    # Sheet 7: Price History
    _create_price_history_sheet(wb.create_sheet("Price History"), run_result)

    wb.save(str(xlsx_path))


def _style_header_row(ws, num_cols: int, row: int = 1):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, num_cols: int, alt: bool = False):
    """Apply data row styling."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_ROW_FILL


def _auto_column_widths(ws, min_width: int = 10, max_width: int = 50):
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _safe_str(value) -> str:
    """Safely convert a value to string for display."""
    if value is None:
        return ""
    if isinstance(value, dict):
        return "; ".join(f"{k}: {v}" for k, v in value.items())
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return str(value)


def _create_overview_sheet(ws, run_result: RunResult):
    """Create overview sheet with run metadata and rankings."""
    ws.title = "Overview"

    # Title
    ws['A1'] = 'Australian Property Research Report'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')

    # Run metadata
    metadata = [
        ('Run ID', run_result.run_id),
        ('Date', run_result.timestamp.strftime('%Y-%m-%d %H:%M:%S')),
        ('Provider', run_result.user_input.provider.title()),
        ('Regions', ', '.join(run_result.user_input.regions)),
        ('Dwelling Type', run_result.user_input.dwelling_type.title()),
        ('Max Price', run_result.user_input.max_median_price),
        ('Suburbs Analyzed', len(run_result.suburbs)),
    ]
    for i, (label, value) in enumerate(metadata, start=3):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=i, column=2, value=value)
        if label == 'Max Price':
            cell.number_format = CURRENCY_FORMAT

    # Rankings table
    start_row = 12
    ws.cell(row=start_row - 1, column=1, value='Suburb Rankings').font = Font(
        name='Calibri', bold=True, size=13, color='2E86AB'
    )

    headers = ['Rank', 'Suburb', 'State', 'LGA', 'Region', 'Median Price',
               '5yr Growth %', 'Growth Score', 'Risk Score', 'Composite Score', 'Data Quality']
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, len(headers), start_row)

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = start_row + 1 + i
        m = report.metrics
        ws.cell(row=row, column=1, value=report.rank)
        ws.cell(row=row, column=2, value=m.identification.name)
        ws.cell(row=row, column=3, value=m.identification.state)
        ws.cell(row=row, column=4, value=m.identification.lga)
        ws.cell(row=row, column=5, value=m.identification.region or '')
        price_cell = ws.cell(row=row, column=6, value=m.market_current.median_price)
        price_cell.number_format = CURRENCY_FORMAT
        growth_cell = ws.cell(row=row, column=7, value=m.growth_projections.projected_growth_pct.get(5, 0))
        growth_cell.number_format = PERCENT_FORMAT
        ws.cell(row=row, column=8, value=m.growth_projections.growth_score)
        ws.cell(row=row, column=9, value=m.growth_projections.risk_score)
        ws.cell(row=row, column=10, value=m.growth_projections.composite_score)
        ws.cell(row=row, column=11, value=m.data_quality.upper())
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)


def _create_market_metrics_sheet(ws, run_result: RunResult):
    """Create market metrics sheet."""
    headers = ['Suburb', 'State', 'Median Price', 'Average Price',
               'Auction Clearance %', 'Days on Market', 'Turnover Rate %',
               'Rental Yield %']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        mc = m.market_current
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)
        c = ws.cell(row=row, column=3, value=mc.median_price)
        c.number_format = CURRENCY_FORMAT
        c = ws.cell(row=row, column=4, value=mc.average_price)
        if mc.average_price:
            c.number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=5, value=mc.auction_clearance_current)
        ws.cell(row=row, column=6, value=mc.days_on_market_current)
        ws.cell(row=row, column=7, value=mc.turnover_rate_current)
        ws.cell(row=row, column=8, value=mc.rental_yield_current)
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)


def _create_growth_sheet(ws, run_result: RunResult):
    """Create growth projections sheet with confidence intervals."""
    headers = ['Suburb', 'State',
               '1yr %', '2yr %', '3yr %', '5yr %', '10yr %', '25yr %',
               '1yr Low', '1yr High', '5yr Low', '5yr High',
               '10yr Low', '10yr High',
               'Growth Score', 'Risk Score', 'Composite Score']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        gp = m.growth_projections
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)

        # Growth percentages
        for col_offset, year in enumerate([1, 2, 3, 5, 10, 25]):
            val = gp.projected_growth_pct.get(year, 0)
            c = ws.cell(row=row, column=3 + col_offset, value=val)
            c.number_format = PERCENT_FORMAT

        # Confidence intervals for 1yr, 5yr, 10yr
        ci = gp.confidence_intervals or {}
        col_idx = 9
        for year in [1, 5, 10]:
            interval = ci.get(year)
            if interval and len(interval) == 2:
                ws.cell(row=row, column=col_idx, value=interval[0]).number_format = PERCENT_FORMAT
                ws.cell(row=row, column=col_idx + 1, value=interval[1]).number_format = PERCENT_FORMAT
            col_idx += 2

        # Scores
        ws.cell(row=row, column=15, value=gp.growth_score)
        ws.cell(row=row, column=16, value=gp.risk_score)
        ws.cell(row=row, column=17, value=gp.composite_score)
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)


def _create_property_config_sheet(ws, run_result: RunResult):
    """Create property configuration sheet."""
    headers = ['Suburb', 'State', 'Land Size (sqm)', 'Floor Size (sqm)',
               'Bedrooms', 'Bathrooms', 'Car Spaces']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        pc = m.physical_config
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)
        ws.cell(row=row, column=3, value=pc.land_size_median_sqm)
        ws.cell(row=row, column=4, value=pc.floor_size_median_sqm)
        ws.cell(row=row, column=5, value=pc.typical_bedrooms)
        ws.cell(row=row, column=6, value=pc.typical_bathrooms)
        ws.cell(row=row, column=7, value=pc.typical_car_spaces)
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)


def _create_demographics_sheet(ws, run_result: RunResult):
    """Create demographics sheet with expanded household and income columns."""
    headers = ['Suburb', 'State', 'Median Age', 'Population Trend',
               'Families with Children %', 'Couples %', 'Single Person %',
               'Group Households %', 'Other Households %',
               'Low Income %', 'Medium Income %', 'High Income %']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        d = m.demographics
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)
        ws.cell(row=row, column=3, value=d.median_age)
        ws.cell(row=row, column=4, value=_safe_str(d.population_trend))

        # Household types - extract from dict with multiple possible keys
        household_types = d.household_types or {}
        ws.cell(row=row, column=5, value=household_types.get('families_with_children', household_types.get('families', 0)))
        ws.cell(row=row, column=6, value=household_types.get('couples', household_types.get('couple_only', 0)))
        ws.cell(row=row, column=7, value=household_types.get('single_person', household_types.get('single', 0)))
        ws.cell(row=row, column=8, value=household_types.get('group_households', household_types.get('group', 0)))
        ws.cell(row=row, column=9, value=household_types.get('other', household_types.get('other_households', 0)))

        # Income distribution - extract from dict with multiple possible keys
        income_dist = d.income_distribution or {}
        ws.cell(row=row, column=10, value=income_dist.get('low_income', income_dist.get('low', 0)))
        ws.cell(row=row, column=11, value=income_dist.get('medium_income', income_dist.get('medium', 0)))
        ws.cell(row=row, column=12, value=income_dist.get('high_income', income_dist.get('high', 0)))

        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)


def _create_infrastructure_sheet(ws, run_result: RunResult):
    """Create infrastructure sheet with expanded columns."""
    headers = ['Suburb', 'State', 'Current Transport', 'Future Transport',
               'Current Infrastructure', 'Planned Infrastructure',
               'Major Events Impact', 'Schools Summary', 'Crime Stats Summary', 'Shopping Access']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    def _format_list_field(field_value):
        """Format list fields with newlines instead of semicolons for better Excel readability."""
        if isinstance(field_value, list):
            return "\n".join(str(item) for item in field_value)
        return _safe_str(field_value)

    def _format_dict_field(field_value):
        """Format dict fields as key: value pairs on separate lines."""
        if isinstance(field_value, dict):
            return "\n".join(f"{k}: {v}" for k, v in field_value.items())
        return _safe_str(field_value)

    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        inf = m.infrastructure
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)
        ws.cell(row=row, column=3, value=_format_list_field(inf.current_transport))
        ws.cell(row=row, column=4, value=_format_list_field(inf.future_transport))
        ws.cell(row=row, column=5, value=_format_list_field(inf.current_infrastructure))
        ws.cell(row=row, column=6, value=_format_list_field(inf.planned_infrastructure))
        ws.cell(row=row, column=7, value=_safe_str(inf.major_events_relevance))
        ws.cell(row=row, column=8, value=_safe_str(inf.schools_summary))
        ws.cell(row=row, column=9, value=_format_dict_field(inf.crime_stats))
        ws.cell(row=row, column=10, value=_safe_str(inf.shopping_access))
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    # Wider columns for text-heavy content
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 40
    for col_letter in ['A', 'B']:
        ws.column_dimensions[col_letter].width = 15


def _create_price_history_sheet(ws, run_result: RunResult):
    """Create price history sheet in wide format (suburbs as rows, years as columns)."""
    # Collect all unique years across all suburbs
    all_years: set[int] = set()
    for report in run_result.get_top_suburbs():
        for tp in report.metrics.market_history.price_history:
            all_years.add(tp.year)

    if not all_years:
        ws.cell(row=1, column=1, value="No price history data available")
        return

    sorted_years = sorted(all_years)

    # Headers
    headers = ['Suburb', 'State'] + [str(y) for y in sorted_years]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(headers))

    # Data rows
    for i, report in enumerate(run_result.get_top_suburbs()):
        row = 2 + i
        m = report.metrics
        ws.cell(row=row, column=1, value=m.identification.name)
        ws.cell(row=row, column=2, value=m.identification.state)

        # Build year -> price lookup
        price_by_year = {tp.year: tp.value for tp in m.market_history.price_history}

        for col_offset, year in enumerate(sorted_years):
            price = price_by_year.get(year)
            if price is not None:
                c = ws.cell(row=row, column=3 + col_offset, value=price)
                c.number_format = CURRENCY_FORMAT

        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    _auto_column_widths(ws)
